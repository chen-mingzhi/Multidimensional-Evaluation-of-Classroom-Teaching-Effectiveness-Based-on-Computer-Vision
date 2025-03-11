import sys
sys.path.append('E:/code/py/github/GFPGAN/FaceRecognize')

import cv2
import joblib
import numpy as np
import pandas as pd
import torch
import warnings
import os
from facenet_pytorch import MTCNN, InceptionResnetV1
from PIL import Image, ImageDraw, ImageFont

import os
import openpyxl

warnings.filterwarnings("ignore", category=UserWarning)


# 人脸识别器
class FaceRecognizer:
    # 初始化，加载数据
    def __init__(self, knn_model_path='knn_model.pkl', face_feature_path='face_feature.csv'):
        # 选择设备
        self.device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')
        # self.device = 'cpu'
        print('Running on device: {}'.format(self.device))

        # 读取训练好的人脸特征数据
        self.data = pd.read_csv(face_feature_path)
        self.x = self.data.drop(columns=['label'])
        self.y = self.data['label']

        # 加载训练好的KNN分类器模型
        self.knn_model = joblib.load(knn_model_path)

        # 字体文件，用于在图片上正确显示中文
        self.font = ImageFont.truetype('simsun.ttc', size=30)

        # mtcnn检测人脸位置
        self.mtcnn = MTCNN(device=self.device, keep_all=True)
        # 用于生成人脸512维特征向量
        self.resnet = InceptionResnetV1(pretrained='vggface2').eval().to(self.device)

    def create_attendance_sheet(self, save_folder_path):
        # 获取E:/code/py/github/GFPGAN/faces下的子目录名字
        dir_path = 'E:/code/py/github/GFPGAN/faces'
        sub_dirs = [d for d in os.listdir(dir_path) if os.path.isdir(os.path.join(dir_path, d))]

        # 创建excel表格
        wb = openpyxl.Workbook()
        sheet = wb.active

        # 设置表头
        sheet['A1'] = 'Name'
        sheet['B1'] = 'Attendance'

        # 写入子目录名字和默认的“未签到”
        for i, sub_dir in enumerate(sub_dirs, 2):
            sheet['A' + str(i)] = sub_dir
            sheet['B' + str(i)] = '未签到'

        # 保存表格
        save_file_path = os.path.join('attendance.xlsx')
        wb.save(save_file_path)
        print("签到记录表格创建成功！保存路径：", save_file_path)

    def update_attendance(self, name, attendance_file, is_attendance=True):
        # 打开excel表格
        wb = openpyxl.load_workbook(attendance_file)
        sheet = wb.active

        # 遍历每行，找到对应的人名进行签到状态更新
        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == name:
                sheet.cell(row=i, column=2, value='已签到' if is_attendance else '未签到')
                break

        # 保存并关闭表格
        wb.save(attendance_file)

    # 人脸识别主函数
    def start_recognize(self, folder_path, save_folder_path, txt_file_path, attendance_file):
    # 创建签到记录表格
        self.create_attendance_sheet(save_folder_path)
        print('--------Face Recognize, start!--------')
        # 获取文件夹中的所有图像文件
        image_files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

        # 如果保存结果的文件夹不存在则创建
        if not os.path.exists(save_folder_path):
            os.makedirs(save_folder_path)

        # 打开txt文件，用于写入人数结果
        txt_file = open(txt_file_path, 'w')

        for image_file in image_files:
            # 读取图像
            image_path = os.path.join(folder_path, image_file)
            image = cv2.imread(image_path)

            img_PIL = Image.fromarray(image)
            draw = ImageDraw.Draw(img_PIL)

            # 检测人脸位置,获得人脸框坐标和人脸概率
            boxes, probs = self.mtcnn.detect(image)
            people_num = len(boxes) if boxes is not None else 0

            # 将人数写入txt文件
            txt_file.write(f"{image_file}: {people_num}\n")

            if boxes is not None:
                for box, prob in zip(boxes, probs):
                    # 设置人脸检测阈值
                    if prob < 0.9:
                        continue

                    x1, y1, x2, y2 = [int(p) for p in box]
                    # 框出人脸位置
                    draw.rectangle((x1, y1, x2, y2), outline=(0, 255, 0), width=2)

                    # 导出人脸图像
                    face = self.mtcnn.extract(image, [box], None).to(self.device)
                    # 生成512维特征向量
                    embeddings = self.resnet(face).detach().cpu().numpy()
                    # KNN预测
                    name_knn = self.knn_model.predict(embeddings)

                    # 获得预测姓名和距离
                    dis = np.linalg.norm(embeddings - self.x.values[:, np.newaxis], axis=2)
                    dis = np.min(dis, axis=0)

                    draw.text((10, 30), f'人数：{people_num}', font=self.font, fill=(255, 255, 0))

                    # 如果距离过大则认为识别失败
                    if dis > 0.65:
                        draw.rectangle((x1, y1, x2, y2), outline=(0, 0, 255), width=2)
                        draw.text((x1, y1 - 40), f'未知', font=self.font, fill=(0, 0, 255))
                    else:
                        # 框出人脸位置并写上名字
                        draw.rectangle((x1, y1, x2, y2), outline=(0, 255, 0), width=2)
                        draw.text((x1, y1 - 40), f'{name_knn[0]}', font=self.font, fill=(0, 255, 0))
                        self.update_attendance(name_knn[0], attendance_file)

                print('检测到人数为:', people_num)
                print('--------Face Recognize, done!--------')

                # 将带有识别结果的图像保存到结果文件夹中
                save_image_path = os.path.join(save_folder_path, f'result_{image_file}')
                img_PIL = np.array(img_PIL)
                cv2.imwrite(save_image_path, img_PIL)
                return people_num

        # 关闭txt文件
        txt_file.close()

        # 打印txt文件的保存路径
        print("人数结果已保存到:", txt_file_path)
